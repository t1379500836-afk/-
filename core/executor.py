"""
工作流执行器 - 在独立线程中执行工作流
"""
import warnings
import threading
import time
from typing import Callable, Optional
import pyautogui
import pyperclip

# 忽略openpyxl的样式警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from .workflow import Workflow, ActionNode, ActionType


class Executor:
    """工作流执行器"""

    def __init__(self):
        self._thread: Optional[threading.Thread] = None
        self._running: bool = False
        self._paused: bool = False
        self._stopped: bool = False
        self._pause_event = threading.Event()
        self._pause_event.set()  # 默认不暂停
        self._on_progress: Optional[Callable] = None
        self._on_complete: Optional[Callable] = None
        self._on_error: Optional[Callable] = None
        self._excel_data = []  # Excel数据
        self._current_loop = 1  # 当前循环次数
        pyautogui.FAILSAFE = True  # 启用安全停止

    @property
    def is_running(self) -> bool:
        return self._running

    @property
    def is_paused(self) -> bool:
        return self._paused

    def set_callbacks(self, on_progress=None, on_complete=None, on_error=None):
        """设置回调函数"""
        self._on_progress = on_progress
        self._on_complete = on_complete
        self._on_error = on_error

    def _execute_single_action(self, node: ActionNode, step_index: int, total_steps: int) -> bool:
        """执行单个操作"""
        try:
            if not node.enabled:
                return True

            if node.action_type in [ActionType.MOUSE_LEFT, ActionType.MOUSE_RIGHT, ActionType.MOUSE_MIDDLE]:
                self._execute_mouse(node)
            elif node.action_type == ActionType.KEYBOARD:
                self._execute_keyboard(node)
            elif node.action_type == ActionType.KEYBOARD_TEXT:
                self._execute_keyboard_text(node)
            elif node.action_type == ActionType.DELAY:
                time.sleep(node.delay_ms / 1000.0)
                return True

            # 报告进度
            if self._on_progress:
                self._on_progress(step_index, total_steps, node.display_name)

            return True
        except Exception as e:
            if self._on_error:
                self._on_error(str(e))
            return False

    def _execute_mouse(self, node: ActionNode) -> None:
        """执行鼠标操作"""
        x, y = node.x, node.y
        button_map = {
            ActionType.MOUSE_LEFT: 'left',
            ActionType.MOUSE_RIGHT: 'right',
            ActionType.MOUSE_MIDDLE: 'middle'
        }
        button = button_map.get(node.action_type, 'left')

        pyautogui.moveTo(x, y, duration=0.1)
        pyautogui.click(button=button)

    def _execute_keyboard(self, node: ActionNode) -> None:
        """执行键盘操作"""
        if node.modifiers:
            keys = node.modifiers + [node.key]
            pyautogui.hotkey(*keys)
        else:
            pyautogui.press(node.key)

    def _execute_keyboard_text(self, node: ActionNode) -> None:
        """执行文本输入"""
        loop_count = getattr(self, '_current_loop', 1)

        # 优先使用Excel数据
        if node.excel_file and self._excel_data:
            row_index = node.excel_start_row - 1 + (loop_count - 1)

            if row_index < len(self._excel_data):
                row = self._excel_data[row_index]
                if node.excel_col - 1 < len(row):
                    text = str(row[node.excel_col - 1])
                    if text and text != 'None':
                        self._type_text(text)
                        return

        # 使用文本行列表（每次循环取一行）
        if node.text_lines:
            index = (loop_count - 1) % len(node.text_lines)
            text = node.text_lines[index].strip()
            if text:
                self._type_text(text)

    def _type_text(self, text: str) -> None:
        """逐字符输入文本"""
        # 处理转义字符
        text = text.replace('\\n', '\n').replace('\\t', '\t').replace('\\\\', '\\')

        pyperclip.copy(text)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.05)

    def _execute_workflow(self, workflow: Workflow) -> None:
        """在工作线程中执行工作流"""
        total_loops = workflow.get_execution_count()
        current_loop = 0

        while self._running:
            if self._stopped:
                break

            current_loop += 1
            if total_loops > 0 and current_loop > total_loops:
                break

            for i, node in enumerate(workflow.nodes):
                if self._stopped:
                    break

                # 处理暂停
                self._pause_event.wait()

                # 在此处再次检查停止标志，因为可能在wait期间被设置
                if self._stopped:
                    break

                # 设置当前循环次数供文本输入使用
                self._current_loop = current_loop

                if not self._execute_single_action(node, i + 1, len(workflow.nodes)):
                    return  # 执行出错停止

                # 步骤间小延迟
                time.sleep(0.05)

        if self._on_complete and not self._stopped:
            self._on_complete()

        self._running = False

    def start(self, workflow: Workflow) -> bool:
        """启动执行"""
        if self._running:
            return False

        self._running = True
        self._paused = False
        self._stopped = False
        self._pause_event.set()
        self._current_loop = 1

        # 加载Excel数据
        self._load_excel_data(workflow)

        self._thread = threading.Thread(target=self._execute_workflow, args=(workflow,), daemon=True)
        self._thread.start()
        return True

    def _load_excel_data(self, workflow: Workflow) -> None:
        """加载工作流中的Excel数据"""
        self._excel_data = []

        # 收集所有Excel文件路径
        excel_files = {}
        for node in workflow.nodes:
            if node.action_type == ActionType.KEYBOARD_TEXT and node.excel_file:
                if node.excel_file not in excel_files:
                    excel_files[node.excel_file] = node.excel_col

        if not excel_files:
            return

        # 读取第一个Excel文件
        for excel_file in excel_files:
            try:
                from openpyxl import load_workbook
                # 不使用read_only模式，确保能读取所有数据
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    self._excel_data.append(list(row))
                wb.close()
                break  # 只读取第一个文件
            except Exception as e:
                print(f"读取Excel失败: {e}")

    def stop(self) -> None:
        """立即停止"""
        self._stopped = True
        self._running = False
        self._pause_event.set()  # 确保如果暂停则恢复以便线程可以检测到停止
        if self._thread and self._thread.is_alive():
            self._thread.join(timeout=1.0)

    def pause(self) -> None:
        """暂停执行"""
        if self._running and not self._paused:
            self._paused = True
            self._pause_event.clear()

    def resume(self) -> None:
        """继续执行"""
        if self._running and self._paused:
            self._paused = False
            self._pause_event.set()